"""
Utilities/ExcelFunctions.py
"""

from openpyxl import workbook
from openpyxl.reader.excel import load_workbook


class Excel_Function:
    def __init__(self,file_name,sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    # Fetch the total row count from the excel sheet
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    # Fetch the total row count from the excel sheet
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    # Read the data from the excel sheet
    def read_data(self,row_number,column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number,column=column_number).value

    # Write the data into the excel sheet
    def write_data(self,row_number,column_number,data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number,column=column_number).value = data
        workbook.save(self.file)




